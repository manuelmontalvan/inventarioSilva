import io
import pandas as pd
from fastapi.responses import StreamingResponse
from cachetools import TTLCache
import hashlib
import json

# Cache para archivos Excel en memoria: max 100 items, 1 hora de TTL
excel_cache = TTLCache(maxsize=100, ttl=3600)

def make_cache_key_excel(forecast: list[dict], alert_restock: bool) -> str:
    """
    Genera una clave hash para el forecast + alerta
    """
    raw_data = {
        "forecast": forecast,
        "alert_restock": alert_restock,
    }
    json_str = json.dumps(raw_data, sort_keys=True)
    return hashlib.sha256(json_str.encode()).hexdigest()

def create_forecast_excel_multi(models: dict, product: str, brand: str, unit: str, days: int):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    wb = Workbook()
    wb.remove(wb.active)  # Elimina hoja vacía inicial

    # Hoja de resumen
    resumen_sheet = wb.create_sheet("Resumen General")
    resumen_sheet.append(["Producto", product])
    resumen_sheet.append(["Marca", brand])
    resumen_sheet.append(["Unidad", unit])
    resumen_sheet.append(["Días de proyección", days])
    resumen_sheet.append([])

    resumen_sheet.append([
        "Modelo",
        "Tendencia",
        "¿Reponer?",
        "Stock Actual",
        "Proyección Total",
        "Variación (%)",
        "MAE",
        "RMSE"
    ])

    # Para resumen gráfico de totales
    resumen_totales = []

    for model_name, model_data in models.items():
        forecast = model_data["forecast"]
        df = pd.DataFrame(forecast)
        df.rename(columns={"ds": "Fecha", "yhat": "Cantidad Estimada"}, inplace=True)

        # Crear hoja individual
        sheet = wb.create_sheet(title=model_name.capitalize())
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

        # Insertar gráfico de línea
        chart = LineChart()
        chart.title = f"Tendencia - {model_name}"
        chart.y_axis.title = "Cantidad Estimada"
        chart.x_axis.title = "Fecha"

        data = Reference(sheet, min_col=2, min_row=1, max_row=len(df) + 1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(df) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True  # Muestra valores en la gráfica

        sheet.add_chart(chart, "D10")

        # Totales y métricas
        percent = model_data.get("percent_change")
        percent = percent if percent is not None else 0

        projected = model_data.get("projected_sales", 0) or 0
        total_projection = int(round(projected))

        metrics = model_data.get("metrics", {})
        mae = metrics.get("MAE")
        rmse = metrics.get("RMSE")

        resumen_sheet.append([
            model_name,
            model_data.get("tendency", ""),
            "Sí" if model_data.get("alert_restock") else "No",
            int(model_data.get("current_quality") or 0),
            total_projection,
            f"{int(round(percent))}%",
            int(round(mae)) if mae is not None else "",
            int(round(rmse)) if rmse is not None else ""
        ])

        resumen_totales.append((model_name, total_projection))

    # Gráfico de barras: Proyección total por modelo
    if resumen_totales:
        chart_sheet = wb.create_sheet("Gráfico Proyecciones")
        chart_sheet.append(["Modelo", "Proyección Total"])

        for name, total in resumen_totales:
            chart_sheet.append([name, total])

        bar_chart = LineChart()
        bar_chart.title = "Total Proyección por Modelo"
        bar_chart.y_axis.title = "Unidades"
        bar_chart.x_axis.title = "Modelo"

        data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(resumen_totales) + 1)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(resumen_totales) + 1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.dLbls = DataLabelList()
        bar_chart.dLbls.showVal = True

        chart_sheet.add_chart(bar_chart, "D5")

    # Guardar archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
